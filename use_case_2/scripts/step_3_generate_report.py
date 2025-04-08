import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from utils.file_handler import read_json_file


def create_excel_from_json(promotion_evaluations_json: str, output_dir: str) -> str:
    """
    creates excel report based on promotion evaluation JSON file

    Args:
    promotion_evaluations_json(str):path to json file created in previous step
    output_dir(str): path to dir where  final excel report will be saved

    return:
    str :path to  final Excel file

    """

    data = read_json_file(promotion_evaluations_json)

    wb = Workbook()
    ws = wb.active
    ws.title = "Evaluation Results"

    headers = [
        "File_name",
        "Facebook_post_text",
        "Promotion_text",
        "Visual_description",
        "Average compliance score",
        "Principle_1_compliance_score",
        "Principle_1_justification",
        "Principle_2_compliance_score",
        "Principle_2_justification",
        "Principle_3_compliance_score",
        "Principle_3_justification",
        "Principle_4_compliance_score",
        "Principle_4_justification",
        "Principle_5_compliance_score",
        "Principle_5_justification",
        "Principle_6_compliance_score",
        "Principle_6_justification",
        "Flagged for supervisory attention",
    ]

    for col_index, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_index, value=header)

    ws.freeze_panes = "A2"

    def compute_flagged(average_score):
        threshold = 2.0
        if average_score < threshold:
            return "Yes"
        else:
            return "No"

    row_num = 2

    for record in data:
        file_name = record.get("image_name", "")
        facebook_post_text = record.get("facebook_post_text", "")
        promotion_text = record.get("promotion_text", "")
        visual_description = record.get("visual_description", "")

        evaluation_results = record.get("evaluation_result", [])

        principles = {}
        for er in evaluation_results:
            p_id = er.get("principle_id")
            p_score = er.get("score", 0)
            p_eval = er.get("justification", "")
            try:
                p_score = float(p_score)
            except ValueError:
                p_score = 0.0
            principles[p_id] = (p_score, p_eval)

        scores = []
        justifications = []
        for pid in ["1", "2", "3", "4", "5", "6"]:
            score, eval_text = principles.get(pid, (0, ""))
            scores.append(score)
            justifications.append(eval_text)

        if scores:
            avg_score = round(sum(scores) / len(scores), 2)
        else:
            avg_score = 0

        flagged = compute_flagged(avg_score)

        ws.cell(row=row_num, column=1, value=file_name)
        ws.cell(row=row_num, column=2, value=facebook_post_text)
        ws.cell(row=row_num, column=3, value=promotion_text)
        ws.cell(row=row_num, column=4, value=visual_description)
        ws.cell(row=row_num, column=5, value=avg_score)

        start_col_for_principles = 6
        for i in range(6):
            score_col = start_col_for_principles + (i * 2)
            just_col = score_col + 1
            ws.cell(row=row_num, column=score_col, value=scores[i])
            ws.cell(row=row_num, column=just_col, value=justifications[i])

        ws.cell(row=row_num, column=18, value=flagged)

        row_num += 1

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{row_num - 1}"

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 50
    ws.column_dimensions["E"].width = 25
    for col_idx in range(6, 18):
        ws.column_dimensions[get_column_letter(col_idx)].width = 25

    flagged_col_index = 18
    flagged_col_letter = get_column_letter(flagged_col_index)
    flagged_range = f"{flagged_col_letter}2:{flagged_col_letter}{row_num - 1}"

    yes_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    no_fill = PatternFill(start_color="008000", end_color="008000", fill_type="solid")
    white_font = Font(color="FFFFFF")

    yes_rule = CellIsRule(
        operator="equal", formula=['"Yes"'], fill=yes_fill, font=white_font
    )
    no_rule = CellIsRule(
        operator="equal", formula=['"No"'], fill=no_fill, font=white_font
    )

    ws.conditional_formatting.add(flagged_range, yes_rule)
    ws.conditional_formatting.add(flagged_range, no_rule)

    ws.sheet_properties.outlinePr.summaryBelow = False
    ws.sheet_properties.outlinePr.summaryRight = True

    ws.column_dimensions.group("F", "Q", outline_level=1, hidden=True)
    consolidated_evaluation_results = os.path.join(
        output_dir, "Consolidated_Evaluation_Results.xlsx"
    )
    wb.save(consolidated_evaluation_results)
    return consolidated_evaluation_results
